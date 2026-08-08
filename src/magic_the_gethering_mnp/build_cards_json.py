import json, re
import openpyxl

wb = openpyxl.load_workbook('mtgnp_master_card_list.xlsx')


# ============================================================
# Structured activated abilities
# ============================================================
#
# The RFC does not define the internal card-catalog schema for
# abilities, so this is an implementation-level structure.
#
# Each ability contains:
#
#   cost:
#       tap  -> whether the permanent must be tapped
#       mana -> mana required to activate
#
#   effect:
#       effect_type -> what happens when the ability resolves
#       additional fields depend on the effect
#
#   targets:
#       whether the ability requires a target
#       what kind of object can be targeted
#
# ============================================================

ACTIVATED_ABILITIES = {

    # --------------------------------------------------------
    # Prodigal Sorcerer
    # Tap: Prodigal Sorcerer deals 1 damage to any target.
    # --------------------------------------------------------
    "prodigal_sorcerer": [
        {
            "cost": {
                "tap": True,
                "mana": {}
            },
            "effect": {
                "effect_type": "DAMAGE",
                "amount": 1
            },
            "targets": {
                "required": True,
                "type": "ANY"
            }
        }
    ],

    # --------------------------------------------------------
    # Rod of Ruin
    # {3}, Tap: Rod of Ruin deals 1 damage to any target.
    # --------------------------------------------------------
    "rod_of_ruin": [
        {
            "cost": {
                "tap": True,
                "mana": {
                    "generic": 3
                }
            },
            "effect": {
                "effect_type": "DAMAGE",
                "amount": 1
            },
            "targets": {
                "required": True,
                "type": "ANY"
            }
        }
    ],

    # --------------------------------------------------------
    # Merfolk Looter
    # Tap: Draw a card, then discard a card.
    # --------------------------------------------------------
    "merfolk_looter": [
        {
            "cost": {
                "tap": True,
                "mana": {}
            },
            "effect": {
                "effect_type": "DRAW_THEN_DISCARD",
                "draw": 1,
                "discard": 1
            },
            "targets": {
                "required": False,
                "type": None
            }
        }
    ],

    # --------------------------------------------------------
    # Millstone
    # {2}, Tap: Target player mills 2 cards.
    # --------------------------------------------------------
    "millstone": [
        {
            "cost": {
                "tap": True,
                "mana": {
                    "generic": 2
                }
            },
            "effect": {
                "effect_type": "MILL",
                "amount": 2
            },
            "targets": {
                "required": True,
                "type": "PLAYER"
            }
        }
    ],

    # --------------------------------------------------------
    # Mother of Runes
    # Tap: Target creature you control gains protection...
    # --------------------------------------------------------
    "mother_of_runes": [
        {
            "cost": {
                "tap": True,
                "mana": {}
            },
            "effect": {
                "effect_type": "GRANT_PROTECTION",
                "duration": "UNTIL_END_OF_TURN"
            },
            "targets": {
                "required": True,
                "type": "CREATURE_YOU_CONTROL"
            }
        }
    ],

    # --------------------------------------------------------
    # Royal Assassin
    # Tap: Destroy target tapped creature.
    # --------------------------------------------------------
    "royal_assassin": [
        {
            "cost": {
                "tap": True,
                "mana": {}
            },
            "effect": {
                "effect_type": "DESTROY"
            },
            "targets": {
                "required": True,
                "type": "TAPPED_CREATURE"
            }
        }
    ],
}


# ============================================================
# Master Card List: base card definitions
# ============================================================

ws_master = wb['Master Card List']
master_rows = list(ws_master.iter_rows(min_row=3, values_only=True))

base_cards = {}

for row in master_rows:
    (
        base_id,
        name,
        ctype,
        subtype,
        color,
        cmc,
        w,
        u,
        b,
        r,
        g,
        generic,
        power,
        toughness,
        copies,
        effect
    ) = row

    mana_cost = {}

    for sym, val in (
        ("W", w),
        ("U", u),
        ("B", b),
        ("R", r),
        ("G", g),
        ("generic", generic)
    ):
        if val:
            mana_cost[sym] = int(val)

    def norm_pt(v):
        if v in ("-", None):
            return None

        try:
            return int(v)
        except (TypeError, ValueError):
            return v

    base_cards[base_id] = {
        "base_id": base_id,
        "name": name,
        "card_type": ctype,
        "subtype": subtype,
        "color": color,
        "cmc": int(cmc) if cmc is not None else 0,
        "mana_cost": mana_cost,
        "power": norm_pt(power),
        "toughness": norm_pt(toughness),
        "copies_in_set": int(copies) if copies is not None else None,
        "effect": effect,

        # ----------------------------------------------------
        # Add structured activated abilities.
        #
        # Cards without an entry receive an empty list.
        # ----------------------------------------------------
        "abilities": ACTIVATED_ABILITIES.get(base_id, []),
    }


# ============================================================
# Card Instances: individual protocol card_id values
# ============================================================

ws_inst = wb['Card Instances']
inst_rows = list(ws_inst.iter_rows(min_row=3, values_only=True))

cards = {}

for row in inst_rows:
    card_id, name, ctype, color, copy_num = row

    base_id = re.sub(r'_\d+$', '', card_id)

    base = base_cards.get(base_id)

    if base is None:
        continue

    entry = dict(base)

    entry["card_id"] = card_id
    entry["copy_number"] = (
        int(copy_num)
        if copy_num is not None
        else None
    )

    del entry["copies_in_set"]

    cards[card_id] = entry


# ============================================================
# Write cards.json
# ============================================================

with open('cards.json', 'w') as f:
    json.dump(cards, f, indent=2, sort_keys=True)

